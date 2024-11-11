import csv
import openpyxl
from openpyxl.styles import PatternFill, Alignment

def normalize_node_name(name):
    return name.replace(' ', '').lower()

def read_hierarchy_csv(filename):
    hierarchy = set()
    max_depth = 0
    with open(filename, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        headers = next(reader, None)  # Skip header if present
        num_levels = len(headers)
        current_path = [''] * num_levels
        for row in reader:
            # Find the first non-empty cell
            for idx, cell in enumerate(row):
                cell = cell.strip()
                if cell:
                    # Update current path at this level
                    current_path[idx] = cell
                    # Clear deeper levels
                    for i in range(idx + 1, num_levels):
                        current_path[i] = ''
                    # Build the path up to this level
                    path = tuple(node for node in current_path if node)
                    hierarchy.add(path)
                    if len(path) > max_depth:
                        max_depth = len(path)
                    break  # Move to the next row after processing
        return [list(path) for path in hierarchy], max_depth

def build_tree(paths):
    tree = {}
    for path in paths:
        current_level = tree
        for node in path:
            current_level = current_level.setdefault(node, {})
    return tree

def compare_trees(base_tree, new_tree, depth=0):
    comparison_rows = []
    base_nodes = {}
    new_nodes = {}

    # Build normalized node mapping for base_tree
    for name in base_tree.keys():
        norm_name = normalize_node_name(name)
        base_nodes[norm_name] = name

    # Similarly for new_tree
    for name in new_tree.keys():
        norm_name = normalize_node_name(name)
        new_nodes[norm_name] = name

    all_norm_names = set(base_nodes.keys()) | set(new_nodes.keys())

    for norm_name in sorted(all_norm_names):
        base_name = base_nodes.get(norm_name)
        new_name = new_nodes.get(norm_name)

        base_subtree = base_tree.get(base_name, {})
        new_subtree = new_tree.get(new_name, {})

        if base_name and new_name:
            status = 'match'
        elif base_name:
            status = 'only_in_base'
        else:
            status = 'only_in_new'

        comparison_rows.append({
            'depth': depth,
            'base_name': base_name,
            'new_name': new_name,
            'status': status
        })

        # Recurse into subtrees
        comparison_rows.extend(compare_trees(base_subtree, new_subtree, depth + 1))

    return comparison_rows

def output_to_excel(comparison_rows, max_depth, base_filename, new_filename, output_filename='comparison_result.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hierarchy Comparison'

    # Define cell styles
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Matching entities
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Only in base
    blue_fill = PatternFill(start_color='CCF2FF', end_color='CCF2FF', fill_type='solid')    # Only in new

    # Total columns = max_depth * 2 + 2 (for two gap columns)
    total_columns = max_depth * 2 + 2

    # Write the file names as the top row
    file_headers = [base_filename] + [''] * (max_depth - 1) + ['', ''] + [new_filename] + [''] * (max_depth - 1)
    ws.append(file_headers)

    # Write header with level numbers only
    headers = [str(i+1) for i in range(max_depth)] + ['', ''] + [str(i+1) for i in range(max_depth)]
    ws.append(headers)

    # Set column widths to minimal
    narrow_width = 25  # Adjusted for better visibility
    gap_width = 3     # Gap column width

    for i in range(1, total_columns + 1):
        column_letter = openpyxl.utils.get_column_letter(i)
        if i == max_depth + 1 or i == max_depth + 2:
            # Gap columns
            ws.column_dimensions[column_letter].width = gap_width
        else:
            ws.column_dimensions[column_letter].width = narrow_width

    # Disable text wrapping for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    # Write comparison rows
    for row_data in comparison_rows:
        depth = row_data['depth']
        base_name = row_data.get('base_name')
        new_name = row_data.get('new_name')
        status = row_data.get('status')

        # Initialize the row with empty cells
        row = [''] * total_columns

        # Base hierarchy columns
        if base_name:
            row[depth] = base_name  # Index from 0
        # New hierarchy columns
        if new_name:
            row[max_depth + 2 + depth] = new_name  # +2 for the gap columns

        # Append the row to the worksheet
        ws.append(row)

        # Apply styles
        excel_row_num = ws.max_row

        # Apply fill to the entire row, excluding the gap columns
        for col in range(1, total_columns + 1):
            if col == max_depth + 1 or col == max_depth + 2:
                continue  # Skip gap columns
            cell = ws.cell(row=excel_row_num, column=col)
            if status == 'only_in_base':
                cell.fill = yellow_fill
            elif status == 'only_in_new':
                cell.fill = blue_fill
            elif status == 'match':
                cell.fill = green_fill
            cell.alignment = Alignment(wrap_text=False)

    wb.save(output_filename)

def main():
    # Replace with your actual file names
    base_hierarchy_csv = 'cco1.csv'  # Reference ontology
    new_hierarchy_csv = 'core1.csv'  # Comparing ontology

    # Read the hierarchies from CSV files
    base_paths, base_max_depth = read_hierarchy_csv(base_hierarchy_csv)
    new_paths, new_max_depth = read_hierarchy_csv(new_hierarchy_csv)

    # Build trees from paths
    base_tree = build_tree(base_paths)
    new_tree = build_tree(new_paths)

    # Determine the maximum depth across both hierarchies
    max_depth = max(base_max_depth, new_max_depth)

    # Compare the trees
    comparison_rows = compare_trees(base_tree, new_tree)

    # Output to Excel with file names
    output_to_excel(comparison_rows, max_depth, base_hierarchy_csv, new_hierarchy_csv)

if __name__ == '__main__':
    main()



============


import csv

def hierarchy_to_csv(input_filename, output_filename):
    rows = []
    max_depth = 0

    with open(input_filename, 'r', encoding='utf-8') as infile:
        lines = infile.readlines()

    for line_num, line in enumerate(lines, start=1):
        # Remove newline character and replace tabs with 4 spaces
        line = line.rstrip('\n').replace('\t', '    ')
        # Skip empty lines
        if not line.strip():
            continue
        # Determine the indentation level
        leading_whitespace = len(line) - len(line.lstrip(' '))
        level = leading_whitespace // 4  # Assuming 4 spaces per indent level
        name = line.strip()
        # Build the row
        row = [''] * level + [name]
        # Update max_depth if necessary
        if len(row) > max_depth:
            max_depth = len(row)
        rows.append(row)

    # Adjust all rows to have the same length
    adjusted_rows = []
    for row in rows:
        row += [''] * (max_depth - len(row))
        adjusted_rows.append(row)

    # Prepare header
    header = ['Level{}'.format(i+1) for i in range(max_depth)]

    # Write to CSV
    with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)
        writer.writerows(adjusted_rows)

if __name__ == '__main__':
    hierarchy_to_csv('hi31.txt', 'hi34.csv')
